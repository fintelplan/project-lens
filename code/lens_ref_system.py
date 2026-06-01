"""
lens_ref_system.py v3
Project Lens — Article Reference System

LENS-022: Redesigned to reflect S1/S2 architecture.

S1 mode (--mode s1):
  Sheet 1 — Articles from STATE/TIER1/TIER2/TIER3 sources (S1's collection pool)
  Sheet 2 — Articles S1 actually scored (from lens_reports.articles_used JSON)
  Filename: YYYYMMDD_S1_1of2.xlsx / YYYYMMDD_S1_2of2.xlsx

S2 mode (--mode s2):
  Sheet 1 — ALL articles (all tiers) + lens_tiercd_data rows (S2's full pool)
  Sheet 2 — Articles S2 flagged for injection (from injection_reports)
            Column: also_s1_pool = YES if STATE/TIER1/TIER2/TIER3, NO if TIER_B/TIER_A/TierCD
  Filename: YYYYMMDD_S2_1of2.xlsx / YYYYMMDD_S2_2of2.xlsx

Both modes run inside lens-manage-analyze.yml after S2 completes.
lens-ref-free.yml and lens-ref-sonnet.yml are deleted (LENS-022).

Usage:
  python code/lens_ref_system.py --mode s1
  python code/lens_ref_system.py --mode s2
"""

import os, json, logging, sys, tempfile
from datetime import datetime, timezone, timedelta

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [REF-SYS] %(levelname)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("REF_SYS")

# S1 source tiers — canary pool
S1_TIERS = {"STATE", "TIER1", "TIER2", "TIER3"}
# S2 extra tiers — beyond S1
S2_EXTRA_TIERS = {"TIER_B", "TIER_A"}


def get_supabase():
    from supabase import create_client
    return create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_KEY"])


def load_source_tier_map() -> dict:
    """Returns {source_id: tier} from lens-SRC-001_sources.json."""
    paths = [
        "data/lens-SRC-001_sources.json",
        os.path.join(os.path.dirname(__file__), "..", "data", "lens-SRC-001_sources.json"),
    ]
    for p in paths:
        if os.path.exists(p):
            with open(p, encoding="utf-8") as f:
                data = json.load(f)
            sources = data.get("sources", data) if isinstance(data, dict) else data
            return {s["id"]: s.get("tier", "TIER2") for s in sources if "id" in s}
    log.warning("lens-SRC-001_sources.json not found — all sources treated as TIER2")
    return {}


def get_slot() -> str:
    """1of2 or 2of2 based on UTC hour (matches manage-analyze cron: 01:28 + 13:28)."""
    hour = datetime.now(timezone.utc).hour
    return "1of2" if hour < 12 else "2of2"


# ── Assign REF numbers ────────────────────────────────────────────────────────

def assign_refs(sb, hours_back: int = 6) -> list:
    """Assign REF-YYYYMMDD-NNNN to new articles. Returns all refs for today."""
    today     = datetime.now(timezone.utc).strftime("%Y%m%d")
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    cutoff    = (datetime.now(timezone.utc) - timedelta(hours=hours_back)).isoformat()

    try:
        r = sb.table("lens_raw_articles") \
            .select("id,title,url,source_id,source_name,domain,collected_at") \
            .gte("collected_at", cutoff) \
            .order("collected_at", desc=False).execute()
        articles = r.data or []
    except Exception as e:
        log.error(f"Fetch articles failed: {e}")
        return []

    if not articles:
        log.info("No articles in window")
        return []

    try:
        ex = sb.table("lens_article_refs") \
            .select("raw_article_id") \
            .eq("collected_date", today_iso).execute()
        existing_ids = {r["raw_article_id"] for r in (ex.data or [])}
    except Exception:
        existing_ids = set()

    new_articles = [a for a in articles if a.get("id") not in existing_ids]

    if not new_articles:
        log.info("All articles already have REF IDs — returning existing")
        try:
            r2 = sb.table("lens_article_refs") \
                .select("ref_id,collected_date,domain,source_name,title,url,raw_article_id") \
                .eq("collected_date", today_iso) \
                .order("ref_id", desc=False).execute()
            return r2.data or []
        except Exception:
            return []

    try:
        mx = sb.table("lens_article_refs") \
            .select("ref_id") \
            .like("ref_id", f"REF-{today}-%") \
            .order("ref_id", desc=True).limit(1).execute()
        last_seq = int(mx.data[0]["ref_id"].split("-")[-1]) if mx.data else 0
    except Exception:
        last_seq = 0

    rows = []
    for i, art in enumerate(new_articles):
        rows.append({
            "ref_id":         f"REF-{today}-{last_seq + i + 1:04d}",
            "collected_date": today_iso,
            "domain":         (art.get("domain") or "GENERAL").upper(),
            "source_name":    art.get("source_name") or "Unknown",
            "title":          (art.get("title") or "")[:300],
            "url":            art.get("url", ""),
            "raw_article_id": art.get("id"),
            # source_id kept in memory only — not in DB schema
            "_source_id":     art.get("source_id", ""),
        })

    # Upsert rows — strip in-memory only fields before DB write
    db_rows = [{k: v for k, v in r.items() if not k.startswith("_")} for r in rows]
    for i in range(0, len(db_rows), 50):
        try:
            sb.table("lens_article_refs").upsert(
                db_rows[i:i+50], on_conflict="raw_article_id").execute()
        except Exception as e:
            log.error(f"Insert batch failed: {e}")

    log.info(f"Referenced {len(new_articles)} new articles")

    # Return in-memory rows (includes _source_id for tier lookup this run)
    try:
        r3 = sb.table("lens_article_refs") \
            .select("ref_id,collected_date,domain,source_name,title,url,raw_article_id") \
            .eq("collected_date", today_iso) \
            .order("ref_id", desc=False).execute()
        # Re-attach _source_id from in-memory rows map
        sid_map = {r["raw_article_id"]: r.get("_source_id", "") for r in rows}
        enriched = []
        for ref in (r3.data or []):
            ref["source_id"] = sid_map.get(ref.get("raw_article_id", ""), "")
            enriched.append(ref)
        return enriched
    except Exception:
        # Fallback: return in-memory rows with _source_id renamed
        for r in rows:
            r["source_id"] = r.pop("_source_id", "")
        return rows


# ── S1: get articles S1 actually scored ──────────────────────────────────────

def get_s1_selected(sb, all_refs: list, hours_back: int = 6) -> list:
    """
    Articles S1 actually scored — parsed from lens_reports.articles_used JSON.
    Joins back to all_refs by raw article id or url.
    Deduplicates; if same article selected by multiple domains, marks with
    selected_by_domains = comma-separated list.
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours_back)).isoformat()

    try:
        r = sb.table("lens_reports") \
            .select("system,domain_focus,articles_used,generated_at") \
            .gte("generated_at", cutoff) \
            .eq("system", "S1") \
            .execute()
        reports = r.data or []
    except Exception as e:
        log.warning(f"lens_reports fetch failed: {e}")
        reports = []

    if not reports:
        log.info("No S1 reports in window")
        return []

    # Build ref lookup maps
    ref_by_id  = {r["raw_article_id"]: r for r in all_refs if r.get("raw_article_id")}
    ref_by_url = {r["url"]: r for r in all_refs if r.get("url")}

    # Collect article_id → set of domains that selected it
    article_domains: dict = {}

    for report in reports:
        domain = report.get("domain_focus", "ALL")
        raw = report.get("articles_used")
        if not raw:
            continue
        try:
            parsed = json.loads(raw) if isinstance(raw, str) else raw
            selected = parsed.get("selected", []) if isinstance(parsed, dict) else []
        except Exception:
            continue

        for art in selected:
            art_id  = art.get("id", "")
            art_url = art.get("url", "")
            key = art_id or art_url
            if not key:
                continue
            if key not in article_domains:
                article_domains[key] = {"art": art, "domains": set()}
            article_domains[key]["domains"].add(domain)

    # Build selected list with duplicate marking
    selected_out = []
    for key, val in article_domains.items():
        art     = val["art"]
        domains = sorted(val["domains"])
        art_id  = art.get("id", "")
        art_url = art.get("url", "")

        ref = ref_by_id.get(art_id) or ref_by_url.get(art_url)
        if ref:
            row = dict(ref)
        else:
            row = {
                "ref_id":       "NO-REF",
                "domain":       art.get("domain", "GENERAL"),
                "source_name":  "Unknown",
                "title":        art.get("title", "")[:300],
                "url":          art_url,
                "raw_article_id": art_id,
            }

        row["selected_by_domains"] = ", ".join(domains)
        row["duplicate_flag"] = "MULTI-DOMAIN" if len(domains) > 1 else ""
        selected_out.append(row)

    log.info(f"S1 selected: {len(selected_out)} articles "
             f"({sum(1 for r in selected_out if r['duplicate_flag'])} multi-domain)")
    return selected_out


# ── S2: get articles S2 flagged ───────────────────────────────────────────────

def get_s2_selected(sb, all_refs: list, tier_map: dict, hours_back: int = 6) -> list:
    """
    Articles S2 flagged for injection — from injection_reports.
    Marks also_s1_pool = YES/NO based on source tier.
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours_back)).isoformat()

    try:
        r = sb.table("injection_reports") \
            .select("analyst,injection_type,confidence_score,flagged_phrases,evidence,created_at") \
            .gte("created_at", cutoff) \
            .order("confidence_score", desc=True).execute()
        injections = r.data or []
    except Exception as e:
        log.warning(f"injection_reports fetch failed: {e}")
        injections = []

    adversarial_sources = {
        "TASS", "RT", "Kremlin", "Global Times", "Xinhua", "CGTN",
        "PressTV", "Press TV", "Iran Press", "Tasnim News",
        "Sputnik", "RIA Novosti", "Dawn", "Asia Times",
        "Middle East Eye", "The Hindu",
    }

    flagged_phrases = set()
    for inj in injections:
        ph = inj.get("flagged_phrases")
        if ph:
            if isinstance(ph, list):
                flagged_phrases.update(str(p).lower() for p in ph if p)
            elif isinstance(ph, str):
                try:
                    parsed = json.loads(ph)
                    if isinstance(parsed, list):
                        flagged_phrases.update(str(p).lower() for p in parsed if p)
                except Exception:
                    flagged_phrases.add(ph.lower())

    analyst_map = {inj.get("analyst"): inj for inj in injections}
    selected = []
    seen_refs = set()

    for ref in all_refs:
        ref_id  = ref.get("ref_id", "")
        source  = ref.get("source_name", "")
        title   = (ref.get("title", "") or "").lower()
        src_id  = ref.get("source_id", "")
        tier    = tier_map.get(src_id, "TIER2")

        also_s1 = "YES" if tier in S1_TIERS else "NO"
        reason  = None
        s2_pos  = None
        finding = None

        if ref_id in seen_refs:
            continue

        if source in adversarial_sources:
            reason  = "Adversarial source"
            s2_pos  = "S2-D"
            inj     = analyst_map.get("S2-D")
            finding = inj.get("injection_type", "") if inj else "Adversary narrative"

        elif flagged_phrases:
            for phrase in flagged_phrases:
                if phrase and len(phrase) > 3 and phrase in title:
                    reason  = f"Flagged phrase: {phrase}"
                    s2_pos  = "S2-A"
                    inj     = analyst_map.get("S2-A")
                    finding = inj.get("injection_type", "") if inj else "Phrase sync"
                    break

        if reason:
            seen_refs.add(ref_id)
            selected.append({
                **ref,
                "s2_position":  s2_pos,
                "finding":      finding,
                "reason":       reason,
                "also_s1_pool": also_s1,
            })

    log.info(f"S2 selected: {len(selected)} articles flagged")
    return selected


# ── Build Excel ───────────────────────────────────────────────────────────────

def build_excel(sheet1_rows: list, sheet2_rows: list,
                mode: str, slot: str, date_str: str, filename: str) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    domain_colors = {
        "MILITARY":  "FFE0E0", "FINANCE":   "E0F0FF",
        "NARRATIVE": "FFF0E0", "POWER":     "F0E0FF",
        "TECH":      "E0FFE0", "NETWORK":   "E0FFFF",
        "RESOURCE":  "FFFDE0", "GENERAL":   "F5F5F5",
    }
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    def make_sheet(ws, rows, col_specs):
        """col_specs = [(header, data_key, width), ...]"""
        headers   = [c[0] for c in col_specs]
        keys      = [c[1] for c in col_specs]
        widths    = [c[2] for c in col_specs]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[ws.cell(1, col).column_letter].width = w
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

        for row_i, row_data in enumerate(rows, 2):
            domain     = (row_data.get("domain") or "GENERAL").upper()
            fill_color = domain_colors.get(domain, "FFFFFF")
            row_fill   = PatternFill(start_color=fill_color,
                                     end_color=fill_color, fill_type="solid")
            for col_i, key in enumerate(keys, 1):
                val  = row_data.get(key, "")
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center")
                if key == "url" and val:
                    cell.hyperlink = val
                    cell.font = Font(color="0563C1", underline="single")

    # ── Sheet 1 ───────────────────────────────────────────────────────────────
    ws1 = wb.active
    if mode == "s1":
        ws1.title = "S1 Collection Pool"
        make_sheet(ws1, sheet1_rows, [
            ("REF ID",      "ref_id",      22),
            ("Domain",      "domain",      13),
            ("Date",        "collected_date", 13),
            ("Source",      "source_name", 22),
            ("Title",       "title",       55),
            ("URL",         "url",         55),
        ])
    else:
        ws1.title = "S2 Full Pool (All Sources)"
        make_sheet(ws1, sheet1_rows, [
            ("REF ID",      "ref_id",      22),
            ("Domain",      "domain",      13),
            ("Date",        "collected_date", 13),
            ("Source",      "source_name", 22),
            ("Tier",        "source_tier", 10),
            ("Also S1?",    "also_s1_pool", 10),
            ("Title",       "title",       45),
            ("URL",         "url",         45),
        ])

    # ── Sheet 2 ───────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet()
    if mode == "s1":
        ws2.title = "S1 Scored Articles"
        make_sheet(ws2, sheet2_rows, [
            ("REF ID",           "ref_id",             22),
            ("Domain",           "domain",             13),
            ("Date",             "collected_date",     13),
            ("Source",           "source_name",        22),
            ("Selected By",      "selected_by_domains",18),
            ("Duplicate Flag",   "duplicate_flag",     16),
            ("Title",            "title",              45),
            ("URL",              "url",                45),
        ])
    else:
        ws2.title = "S2 Flagged Articles"
        make_sheet(ws2, sheet2_rows, [
            ("REF ID",      "ref_id",      22),
            ("Domain",      "domain",      13),
            ("Date",        "collected_date", 13),
            ("Source",      "source_name", 20),
            ("Also S1?",    "also_s1_pool", 10),
            ("S2 Position", "s2_position", 14),
            ("Finding",     "finding",     25),
            ("Reason",      "reason",      30),
            ("Title",       "title",       40),
            ("URL",         "url",         40),
        ])

    # ── Sheet 3 Summary ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Project Lens — Article References"
    ws3["A2"] = f"System: {'S1 (Canary Pool)' if mode == 's1' else 'S2 (Full Pool)'}"
    ws3["A3"] = f"Date: {date_str}"
    ws3["A4"] = f"Slot: {slot}"
    ws3["A5"] = f"Sheet 1 (Collection pool): {len(sheet1_rows)}"
    ws3["A6"] = f"Sheet 2 (Selected/Flagged): {len(sheet2_rows)}"

    domain_counts = {}
    for ref in sheet1_rows:
        d = ref.get("domain", "GENERAL")
        domain_counts[d] = domain_counts.get(d, 0) + 1
    ws3["A8"] = "By Domain:"
    for i, (d, c) in enumerate(sorted(domain_counts.items(), key=lambda x: -x[1]), 9):
        ws3[f"A{i}"] = d
        ws3[f"B{i}"] = c

    if mode == "s2":
        tier_counts = {}
        for ref in sheet1_rows:
            t = ref.get("source_tier", "?")
            tier_counts[t] = tier_counts.get(t, 0) + 1
        ws3["D8"] = "By Tier:"
        for i, (t, c) in enumerate(sorted(tier_counts.items(), key=lambda x: -x[1]), 9):
            ws3[f"D{i}"] = t
            ws3[f"E{i}"] = c

    # -- Sheet 4: Unflagged Titles with Links (pool minus scored/flagged) --
    scored_flagged_ids = {r.get("ref_id") for r in sheet2_rows if r.get("ref_id")}
    unflagged_rows = [r for r in sheet1_rows
                      if r.get("ref_id") and r.get("ref_id") not in scored_flagged_ids]
    ws4 = wb.create_sheet("Unflagged Titles with Links")
    make_sheet(ws4, unflagged_rows, [
        ("REF ID",  "ref_id",      22),
        ("Source",  "source_name", 24),
        ("Title",   "title",       60),
        ("URL",     "url",         60),
    ])
    ws3["A7"] = f"Sheet 4 (Unflagged titles+links): {len(unflagged_rows)}"

    out = os.path.join(tempfile.gettempdir(), filename)
    wb.save(out)
    log.info(f"Excel saved: {out}")
    return out


# ── Telegram send ─────────────────────────────────────────────────────────────

def send_telegram(path: str, filename: str, mode: str, slot: str,
                  total: int, selected: int, date_str: str) -> bool:
    import requests
    token   = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")
    if not token or not chat_id:
        return False
    try:
        sys_label = "S1 Canary Pool" if mode == "s1" else "S2 Full Pool"
        sel_label = "Scored by S1" if mode == "s1" else "Flagged by S2"
        caption = (
            f"📊 {filename}\n"
            f"{date_str} | {sys_label} | {slot}\n\n"
            f"Sheet 1 (Collection): {total}\n"
            f"Sheet 2 ({sel_label}): {selected}\n\n"
            f"Sheet 1: Collection Pool\n"
            f"Sheet 2: {'Scored Articles' if mode == 's1' else 'Flagged Articles'}\n"
            f"Sheet 3: Summary\n"
            f"Sheet 4: Unflagged Titles with Links"
        )
        url = f"https://api.telegram.org/bot{token}/sendDocument"
        with open(path, "rb") as f:
            resp = requests.post(url,
                data={"chat_id": chat_id, "caption": caption},
                files={"document": (filename, f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=60)
        ok = resp.status_code == 200
        log.info(f"Telegram: {'OK' if ok else 'FAILED'}")
        return ok
    except Exception as e:
        log.error(f"Telegram failed: {e}")
        return False


# ── Main ──────────────────────────────────────────────────────────────────────

def run(mode: str = "s1") -> dict:
    import time
    start    = time.time()
    slot     = get_slot()
    today    = datetime.now(timezone.utc).strftime("%Y%m%d")
    thai     = datetime.now(timezone.utc) + timedelta(hours=7)
    date_str = thai.strftime("%B %d, %Y %I:%M %p")
    filename = f"{today}_{'S1' if mode == 's1' else 'S2'}_{slot}.xlsx"

    log.info(f"=== REF SYSTEM START | mode={mode} slot={slot} file={filename} ===")

    try:
        sb = get_supabase()
    except Exception as e:
        return {"status": "ERROR", "error": str(e)}

    tier_map = load_source_tier_map()

    # Assign REFs (always 6h window)
    all_refs = assign_refs(sb, hours_back=6)
    if not all_refs:
        log.warning("No articles found — skipping export")
        return {"status": "NO_DATA"}

    if mode == "s1":
        # Sheet 1: S1 pool only (STATE/TIER1/TIER2/TIER3)
        sheet1 = [r for r in all_refs
                  if tier_map.get(r.get("source_id", ""), "TIER2") in S1_TIERS]
        # Sheet 2: articles S1 actually scored
        sheet2 = get_s1_selected(sb, all_refs, hours_back=6)

    else:
        # Sheet 1: full S2 pool — all tiers + tiercd_data
        # Tag each ref with tier + also_s1_pool
        sheet1 = []
        for r in all_refs:
            tier    = tier_map.get(r.get("source_id", ""), "TIER2")
            also_s1 = "YES" if tier in S1_TIERS else "NO"
            sheet1.append({**r, "source_tier": tier, "also_s1_pool": also_s1})

        # Append lens_tiercd_data rows
        try:
            cutoff = (datetime.now(timezone.utc) - timedelta(hours=6)).isoformat()
            tc = sb.table("lens_tiercd_data") \
                .select("id,tier,source_name,data_type,title,url,fetch_date") \
                .gte("created_at", cutoff).execute()
            for row in (tc.data or []):
                sheet1.append({
                    "ref_id":        f"TIERCD-{str(row['id'])[:8]}",
                    "collected_date": row.get("fetch_date", ""),
                    "domain":        "GENERAL",
                    "source_id":     "",
                    "source_name":   row.get("source_name", ""),
                    "source_tier":   row.get("tier", "TIER_C"),
                    "also_s1_pool":  "NO",
                    "title":         (row.get("title") or "")[:300],
                    "url":           row.get("url", ""),
                    "raw_article_id": str(row["id"]),
                })
            log.info(f"TierCD rows appended: {len(tc.data or [])}")
        except Exception as e:
            log.warning(f"TierCD fetch failed (non-fatal): {e}")

        # Sheet 2: S2 flagged articles
        sheet2 = get_s2_selected(sb, all_refs, tier_map, hours_back=6)

    xlsx_path = build_excel(sheet1, sheet2, mode, slot, date_str, filename)
    sent      = send_telegram(xlsx_path, filename, mode, slot,
                               len(sheet1), len(sheet2), date_str)

    elapsed = round(time.time() - start, 1)
    log.info(f"=== DONE | sheet1={len(sheet1)} | sheet2={len(sheet2)} | {elapsed}s ===")
    return {
        "status":   "OK",
        "file":     filename,
        "sheet1":   len(sheet1),
        "sheet2":   len(sheet2),
        "sent":     sent,
    }


if __name__ == "__main__":
    from dotenv import load_dotenv
    load_dotenv()
    mode = "s2"
    for arg in sys.argv[1:]:
        if arg in ("--mode", "-m") and sys.argv.index(arg) + 1 < len(sys.argv):
            mode = sys.argv[sys.argv.index(arg) + 1]
        elif arg in ("s1", "s2"):
            mode = arg
    result = run(mode=mode)
    print(json.dumps(result, indent=2))
